import random

from datetime import datetime, timedelta

from openpyxl import load_workbook

from openpyxl.utils import range_boundaries

# Load the Excel file while preserving formatting

file_path = r"C:\Users\hss930284\Tata Technologies\MBSE Team - SAARCONN - SAARCONN\Eliminating_SystemDesk\tests\Harshit_arelements_validation_24_03\COMBINED_AUTOMATION\Default_Files\Default_Input_Excel.xlsx"  # Change this to your actual file path

wb = load_workbook(file_path, keep_vba=True)  # keep_vba=True ensures macros & formatting are preserved

# Function to extract dropdown choices by parsing cell ranges

def get_dropdown_choices(ws):

    dropdowns = {}

    for dv in ws.data_validations.dataValidation:

        if dv.formula1 and dv.type == "list":

            choices = dv.formula1.replace('"', "").split(",")

            for cell_range in dv.sqref:  # Iterate directly, no need to split

                min_col, min_row, max_col, max_row = range_boundaries(str(cell_range))

                for row in range(min_row, max_row + 1):

                    for col in range(min_col, max_col + 1):

                        cell_coord = ws.cell(row=row, column=col).coordinate

                        dropdowns[cell_coord] = choices  # Store dropdown choices for this cell

    return dropdowns

# Function to modify cell values while respecting dropdowns

def modify_value(value, valid_choices=None):

    if valid_choices:  # If the cell has dropdown values, select a random valid value

        return random.choice(valid_choices)

    elif isinstance(value, (int, float)):  # Adjust numbers within a reasonable range

        return value + random.randint(-5, 5)

    elif isinstance(value, str):  # Slight text modifications

        if value.isdigit():  # If the text is a number, modify it numerically

            return str(int(value) + random.randint(-5, 5))

        elif len(value) > 3:  # Shuffle text slightly

            chars = list(value)

            random.shuffle(chars)

            return "".join(chars)

    elif isinstance(value, datetime):  # Adjust dates by a few days

        return value + timedelta(days=random.randint(-3, 3))

    return value  # Return original if no changes are applicable

# Process each sheet

for sheet_name in wb.sheetnames:

    ws = wb[sheet_name]

    dropdowns = get_dropdown_choices(ws)  # Get dropdown values for the sheet

    # Get all possible modifiable cells excluding headers

    modifiable_cells = []

    for row in ws.iter_rows(min_row=2, values_only=False):  # Skip headers

        for cell in row:

            if cell.value is not None:

                modifiable_cells.append(cell)

    # Select 10 random cells to modify (spread across the sheet)

    if modifiable_cells:

        selected_cells = random.sample(modifiable_cells, min(10, len(modifiable_cells)))

        for cell in selected_cells:

            valid_choices = dropdowns.get(cell.coordinate, None)  # Get dropdown choices if applicable

            cell.value = modify_value(cell.value, valid_choices)  # Modify value while keeping formatting

# Save the modified file

modified_file_path = "Modified_Appl5.xlsx"

wb.save(modified_file_path)

print(f"Modified file saved as: {modified_file_path}") 